from ast import Dict, List
from typing import Any
from fastapi import HTTPException


from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.exc import IntegrityError
from sqlmodel import  or_
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from io import BytesIO
from typing import Any, Dict, Tuple

from app.database.models import Class, Enrollment, Group, Subject, SuperUser, User
from app.core.securty import hash_password,verify_password,create_access_token,decode_token


class SuperUserService():
    def __init__(self,session:AsyncSession):
        self.session = session

    async def create_super_user(self,first_name:str,last_name:str,username:str,password:str,telegram_id = None,is_root = False):

        hashed_password = hash_password(password)
        new_user = SuperUser(first_name = first_name,
                                              last_name = last_name,
                                              username = username,
                                              hashed_password = hashed_password,
                                              telegram_id = telegram_id,
                                              is_root = is_root)
        try:
            self.session.add(new_user)
            await self.session.commit()
            await self.session.refresh(new_user)

            return None
        except IntegrityError as e:
            await self.session.rollback()
            raise HTTPException(detail="Username already exists!",status_code=409)


    async def authenticate_user(self,username:str,password:str)->SuperUser:
        stmt = select(SuperUser).where(SuperUser.username == username)
        query = await self.session.execute(stmt)
        user = query.scalar_one_or_none()

        if not user:
            raise HTTPException(detail="User not found",status_code=404)
        if not verify_password(plain_password=password,hashed_password=user.hashed_password):
            raise HTTPException(detail="Password is incorrect",status_code=403)
        

        access_token = create_access_token({
            "sub":str(user.id)
        })

        return {
            "access_token":access_token,
            "token_type":"bearer"
        }
    



    async def get_super_users(self):
        stmt = await self.session.execute(
            select(SuperUser)
        )
        return stmt.scalars().all()
    
    async def get_user_by_id(self,user_id):
        stmt = await self.session.execute(
            select(SuperUser).where(SuperUser.id == UUID(user_id))
        )
        user = stmt.scalar_one_or_none()
        if not user:
            raise HTTPException(detail="User not found",status_code=404)
        
        return user
    
    async def delete_super_user(self,user_id,current_super_user:SuperUser):
        stmt = await self.session.execute(
            select(SuperUser).where(SuperUser.id == user_id)
        )
        user = stmt.scalar_one_or_none()
        if not user:
            raise HTTPException(detail="User not found",status_code=404)
        
        if user.id == current_super_user.id:
            raise HTTPException(detail="You can not delete your self",status_code=400)
        
        try:
            self.session.delete(user)          
            await self.session.commit()     
            return {"ok": "success"}
        except IntegrityError as e:
            await self.session.rollback()
            raise HTTPException(detail="Database error")
        
    










    async def get_attendance_matrix_by_program_cohort(
        self,
        program,   # GroupType
        cohort: int,
    ):
        """
        Returns an Excel-like matrix for all groups of the given program+cohort,
        ignoring section (e.g. CSE-24-01..CSE-24-16 all included).

        Cell rules:
        - enrolled:  {"status":"enrolled","absence":x,"late":y}
        - dropped:  {"status":"dropped"}   (group has subject, student not enrolled)
        - na:       {"status":"na"}        (group doesn't have that subject)
        """

        program_value = program.value if hasattr(program, "value") else str(program).upper().strip()
        prefix = f"{program_value}-{cohort:02d}"

        # include "CSE-24-12" and also "MBA-25" (no section)
        group_filter = or_(
            Group.group_name == prefix,
            Group.group_name.like(prefix + "-%")
        )

        # 1) Load matching groups
        res_groups = await self.session.execute(select(Group).where(group_filter))
        groups = res_groups.scalars().all()
        group_ids = [g.id for g in groups]
        group_names = sorted([g.group_name for g in groups])

        if not group_ids:
            return {
                "program": program_value,
                "cohort": cohort,
                "groups": [],
                "subjects": [],
                "rows": [],
            }

        # 2) Load all subjects that exist for these groups (Class -> Subject)
        res_gs = await self.session.execute(
            select(Class.group_id, Subject)
            .join(Subject, Subject.id == Class.subject_id)
            .where(Class.group_id.in_(group_ids))
        )
        gs_rows = res_gs.all()  # (group_id, Subject)

        # group_id -> set(subject_id)
        group_subject_ids: Dict[str, set[str]] = {}
        # subject_id -> subject dict (columns)
        subjects_map: Dict[str, Dict[str, Any]] = {}

        for gid, subj in gs_rows:
            gid_str = str(gid)
            sid_str = str(subj.id)

            group_subject_ids.setdefault(gid_str, set()).add(sid_str)
            subjects_map.setdefault(sid_str, {
                "id": sid_str,
                "short_name": subj.short_name,
                "name": subj.name,
            })

        subjects = sorted(
            subjects_map.values(),
            key=lambda x: (x["short_name"] or "", x["name"] or "")
        )
        subject_ids_all = [s["id"] for s in subjects]

        # 3) Load users in those groups
        res_users = await self.session.execute(
            select(User, Group)
            .join(Group, Group.id == User.group_id)
            .where(Group.id.in_(group_ids))
            .where(User.telegram_id.isnot(None))   # ← add this line
        )

        user_rows = res_users.all()  # (User, Group)

        if not user_rows:
            return {
                "program": program_value,
                "cohort": cohort,
                "groups": group_names,
                "subjects": subjects,
                "rows": [],
            }

        users_by_id: Dict[str, Dict[str, Any]] = {}
        user_ids = []
        for user, group in user_rows:
            uid = str(user.id)
            users_by_id[uid] = {"user": user, "group": group}
            user_ids.append(user.id)

        # 4) Load enrollments for those users (Enrollment -> Class gives subject_id & group_id)
        res_enr = await self.session.execute(
            select(
                Enrollment.user_id,
                Class.group_id,
                Class.subject_id,
                Enrollment.absence,
                Enrollment.late,
            )
            .join(Class, Class.id == Enrollment.class_id)
            .where(Enrollment.user_id.in_(user_ids))
            .where(Class.group_id.in_(group_ids))
        )
        enr_rows = res_enr.all()

        # (user_id, subject_id) -> stats
        enr_map: Dict[tuple[str, str], Dict[str, int]] = {}
        for user_id, group_id, subject_id, absence, late in enr_rows:
            enr_map[(str(user_id), str(subject_id))] = {
                "absence": absence or 0,
                "late": late or 0,
            }

        # 5) Build matrix
        rows_out: List[Dict[str, Any]] = []

        for uid, pack in users_by_id.items():
            user = pack["user"]
            group = pack["group"]
            gid_str = str(group.id)

            allowed_subjects = group_subject_ids.get(gid_str, set())

            row = {
                "student": {
                    "id": uid,
                    "student_id": user.student_id,
                    "first_name": user.first_name,
                    "last_name": user.last_name,
                    "group_name": group.group_name,
                },
                "cells": {}
            }

            for sid in subject_ids_all:
                if sid not in allowed_subjects:
                    row["cells"][sid] = {"status": "na"}
                    continue

                stats = enr_map.get((uid, sid))
                if stats is None:
                    row["cells"][sid] = {"status": "dropped"}
                else:
                    row["cells"][sid] = {"status": "enrolled", **stats}

            rows_out.append(row)

        rows_out.sort(key=lambda r: (
            r["student"]["group_name"] or "",
            r["student"]["last_name"] or "",
            r["student"]["first_name"] or "",
        ))

        return {
            "program": program_value,
            "cohort": cohort,
            "groups": group_names,
            "subjects": subjects,
            "rows": rows_out,
        }
        


    #EXCELL

    async def export_attendance_matrix_excel_professional(
        self,
        program,   # GroupType
        cohort: int,
    ) -> Tuple[bytes, str]:
        """
        Creates a professional Excel file with conditional coloring based on absence count.
        Returns: (xlsx_bytes, filename)
        """

        matrix = await self.get_attendance_matrix_by_program_cohort(program, cohort)

        wb = Workbook()
        ws = wb.active
        ws.title = f"{matrix['program']}-{matrix['cohort']}"

        subjects = matrix["subjects"]
        rows = matrix["rows"]

        # ---------- Styles ----------
        title_font = Font(bold=True, size=16, color="FFFFFF")
        header_font = Font(bold=True, color="FFFFFF")
        base_font = Font(size=11, color="111111")

        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Professional palette (soft)
        fill_title = PatternFill("solid", fgColor="0B2F4E")     # dark blue
        fill_header = PatternFill("solid", fgColor="1F4E79")    # blue
        fill_meta = PatternFill("solid", fgColor="E8EEF7")      # very light blue-gray
        fill_na = PatternFill("solid", fgColor="F2F2F2")        # light gray
        fill_dropped = PatternFill("solid", fgColor="D9D9D9")   # gray

        # Absence severity fills
        fill_abs_0 = PatternFill("solid", fgColor="EEF6FF")     # clean pale
        fill_abs_1_2 = PatternFill("solid", fgColor="DCEEFF")   # slightly stronger
        fill_abs_3_4 = PatternFill("solid", fgColor="FFE2C6")   # soft warning
        fill_abs_5_7 = PatternFill("solid", fgColor="FFF2B2")   # yellow warning
        fill_abs_8p = PatternFill("solid", fgColor="FFB3B3")    # red warning

        thin = Side(style="thin", color="A6A6A6")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def absence_fill(absence: int) -> PatternFill:
            if absence <= 0:
                return fill_abs_0
            if 1 <= absence <= 2:
                return fill_abs_1_2
            if 3 <= absence <= 4:
                return fill_abs_3_4
            if 5 <= absence <= 7:
                return fill_abs_5_7
            return fill_abs_8p  # 8+

        # ---------- Layout ----------
        # Row 1: Title merged
        total_cols = 3 + len(subjects)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=f"Attendance Matrix — {matrix['program']}-{matrix['cohort']}")
        title_cell.font = title_font
        title_cell.fill = fill_title
        title_cell.alignment = center
        ws.row_dimensions[1].height = 28

        # Row 2: metadata (optional, looks pro)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        meta_cell = ws.cell(
            row=2,
            column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}    "
                  f"Groups: {len(matrix.get('groups', []))}    Students: {len(rows)}"
        )
        meta_cell.fill = fill_meta
        meta_cell.font = Font(size=10, color="333333")
        meta_cell.alignment = left
        ws.row_dimensions[2].height = 18

        # Row 3: Header
        headers = ["Group", "Student ID", "Full Name"] + [
            (s["short_name"] or s["name"] or "SUBJECT") for s in subjects
        ]
        ws.append([""] * total_cols)  # placeholder to make row index align
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.fill = fill_header
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[3].height = 22

        # Freeze panes below header, enable filters
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(total_cols)}3"

        # Column widths
        ws.column_dimensions["A"].width = 14  # group
        ws.column_dimensions["B"].width = 14  # student id
        ws.column_dimensions["C"].width = 24  # name
        for i in range(4, total_cols + 1):
            ws.column_dimensions[get_column_letter(i)].width = 13

        # ---------- Data ----------
        current_row = 4
        for r in rows:
            student = r["student"]
            cells = r["cells"]

            # Base columns
            ws.cell(row=current_row, column=1, value=student.get("group_name", "")).alignment = center
            ws.cell(row=current_row, column=2, value=student.get("student_id", "")).alignment = center
            ws.cell(
                row=current_row,
                column=3,
                value=f"{student.get('first_name','') or ''} {student.get('last_name','') or ''}".strip()
            ).alignment = left

            # Style base cells (A-C)
            for col in (1, 2, 3):
                c = ws.cell(row=current_row, column=col)
                c.font = base_font
                c.border = border

            # Subject cells
            for idx, subj in enumerate(subjects, start=4):
                sid = subj["id"]
                cell_payload = cells.get(sid, {"status": "na"})
                status = cell_payload.get("status", "na")

                excel_cell = ws.cell(row=current_row, column=idx)
                excel_cell.font = base_font
                excel_cell.alignment = center
                excel_cell.border = border

                if status == "enrolled":
                    a = int(cell_payload.get("absence", 0) or 0)
                    l = int(cell_payload.get("late", 0) or 0)
                    excel_cell.value = f"A:{a}  L:{l}"
                    excel_cell.fill = absence_fill(a)

                elif status == "dropped":
                    excel_cell.value = "Dropped"
                    excel_cell.fill = fill_dropped
                    excel_cell.font = Font(bold=True, color="404040")

                else:  # "na"
                    excel_cell.value = ""
                    excel_cell.fill = fill_na

            ws.row_dimensions[current_row].height = 18
            current_row += 1

        # Optional: make sheet look nice when opened
        ws.sheet_view.showGridLines = False

        # ---------- Save ----------
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{matrix['program']}-{matrix['cohort']}_matrix_{ts}.xlsx"
        return bio.getvalue(), filename


